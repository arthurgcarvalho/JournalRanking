"""
build_data.py — Preprocesses ABDC, AJG, and FSB workload Excel files
into a single journals.json used by the web application.

Uses multi-pass matching to merge journals across ABDC and AJG lists
despite title inconsistencies (subtitles, & vs and, minor spelling).

Usage:
    python build_data.py
"""

import json
import re
import openpyxl
from pathlib import Path

DATA_DIR = Path(__file__).parent / "Data"
OUTPUT = Path(__file__).parent / "journals.json"


# ----------------------------------------------------------------
# Normalization helpers
# ----------------------------------------------------------------

def normalize(title: str) -> str:
    """Basic normalize: lowercase, strip punctuation, collapse spaces."""
    t = title.lower().strip()
    t = re.sub(r"[^a-z0-9\s]", "", t)
    t = re.sub(r"\s+", " ", t)
    return t


def normalize_deep(title: str) -> str:
    """Aggressive normalize: also strip subtitles, parentheticals,
    replace & with and, remove 'the' prefix."""
    t = title
    # Remove parenthetical suffixes like "(United Kingdom)" or "(JAMIS)"
    t = re.sub(r"\s*\(.*?\)\s*", " ", t)
    # Remove subtitle after colon
    t = t.split(":")[0]
    # Replace & with and
    t = t.replace("&", "and")
    # Basic normalize
    t = t.lower().strip()
    t = re.sub(r"[^a-z0-9\s]", "", t)
    t = re.sub(r"\s+", " ", t)
    # Remove leading "the"
    t = re.sub(r"^the\s+", "", t)
    return t


def bigrams(s: str) -> set:
    """Return set of character bigrams."""
    s = s.lower()
    return set(s[i:i+2] for i in range(len(s) - 1))


def dice_coefficient(a: set, b: set) -> float:
    """Dice similarity between two bigram sets."""
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    return (2 * len(a & b)) / (len(a) + len(b))


def clean(val):
    """Strip whitespace from strings, pass through others."""
    if isinstance(val, str):
        return val.strip()
    return val


# ----------------------------------------------------------------
# Data loaders
# ----------------------------------------------------------------

def load_abdc():
    """Load ABDC journal list -> dict keyed by normalized title."""
    wb = openpyxl.load_workbook(DATA_DIR / "ABDC-JQL-2022.xlsx", read_only=True)
    ws = wb.active
    journals = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        title_raw = row[0]
        if not title_raw or not isinstance(title_raw, str):
            continue
        title = clean(title_raw)
        publisher = clean(row[1]) if row[1] else ""
        rating = clean(row[6]) if row[6] else ""
        key = normalize(title)
        journals[key] = {
            "title": title,
            "publisher": publisher,
            "abdc": rating,
        }
    wb.close()
    return journals


def load_ajg():
    """Load AJG journal list -> dict keyed by normalized title."""
    wb = openpyxl.load_workbook(DATA_DIR / "AJG.xlsx", read_only=True)
    ws = wb.active
    journals = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        title_raw = row[1]
        if not title_raw or not isinstance(title_raw, str):
            continue
        title = clean(title_raw)
        field = clean(row[0]) if row[0] else ""
        publisher = clean(row[2]) if row[2] else ""
        ajg_raw = clean(row[3])
        ajg = str(ajg_raw) if ajg_raw is not None else ""
        key = normalize(title)
        journals[key] = {
            "title": title,
            "ajg": ajg,
            "field": field,
            "publisher": publisher,
        }
    wb.close()
    return journals


def load_fsb_points():
    """Load FSB workload point lookup table.
    Returns dict: (abdc_str, ajg_str) -> points.
    """
    wb = openpyxl.load_workbook(DATA_DIR / "FSB workload point system.xlsx", read_only=True)
    ws = wb.active
    table = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        abdc_raw = row[0]
        ajg_raw = row[1]
        points = row[2]
        if points is None:
            continue
        abdc_key = str(clean(abdc_raw)) if abdc_raw is not None else ""
        ajg_key = str(clean(ajg_raw)) if ajg_raw is not None else ""
        table[(abdc_key, ajg_key)] = int(points)
    wb.close()
    return table


def calc_points(abdc: str, ajg: str, table: dict) -> int:
    """Look up FSB points for an (ABDC, AJG) combination.
    Tries exact, then partial fallbacks.
    """
    for key in [(abdc, ajg), (abdc, ""), ("", ajg), ("", "")]:
        if key in table:
            return table[key]
    return 0


# ----------------------------------------------------------------
# Multi-pass journal matching
# ----------------------------------------------------------------

def match_journals(abdc_data: dict, ajg_data: dict):
    """Match AJG journals to ABDC journals using three passes.

    Returns:
        matched: list of (abdc_key, ajg_key) pairs
        unmatched_ajg: set of ajg_keys with no match
        unmatched_abdc: set of abdc_keys with no match
    """
    matched = []
    matched_abdc_keys = set()
    matched_ajg_keys = set()

    # --- Pass 1: Exact normalized key match ---
    for ajg_key in ajg_data:
        if ajg_key in abdc_data:
            matched.append((ajg_key, ajg_key))
            matched_abdc_keys.add(ajg_key)
            matched_ajg_keys.add(ajg_key)

    p1_count = len(matched)

    # --- Pass 2: Deep normalization (strip subtitles, & -> and, etc.) ---
    # Build deep-key lookup for unmatched ABDC entries
    abdc_deep = {}  # deep_key -> original_key
    for abdc_key, entry in abdc_data.items():
        if abdc_key in matched_abdc_keys:
            continue
        dk = normalize_deep(entry["title"])
        if dk and dk not in abdc_deep:
            abdc_deep[dk] = abdc_key

    for ajg_key, entry in ajg_data.items():
        if ajg_key in matched_ajg_keys:
            continue
        dk = normalize_deep(entry["title"])
        if dk in abdc_deep:
            abdc_key = abdc_deep[dk]
            matched.append((abdc_key, ajg_key))
            matched_abdc_keys.add(abdc_key)
            matched_ajg_keys.add(ajg_key)
            # Remove so it can't match again
            del abdc_deep[dk]

    p2_count = len(matched) - p1_count

    # --- Pass 3: Fuzzy bigram matching (dice >= 0.82) ---
    remaining_abdc = {k: v for k, v in abdc_data.items() if k not in matched_abdc_keys}
    remaining_ajg = {k: v for k, v in ajg_data.items() if k not in matched_ajg_keys}

    # Pre-compute deep-normalized keys and bigrams for ABDC
    abdc_fuzzy = []
    for abdc_key, entry in remaining_abdc.items():
        dk = normalize_deep(entry["title"])
        bg = bigrams(dk)
        abdc_fuzzy.append((abdc_key, dk, bg))

    FUZZY_THRESHOLD = 0.82
    fuzzy_candidates = []

    for ajg_key, entry in remaining_ajg.items():
        dk = normalize_deep(entry["title"])
        ajg_bg = bigrams(dk)
        best_score = 0
        best_abdc_key = None
        for abdc_key, adk, abg in abdc_fuzzy:
            score = dice_coefficient(ajg_bg, abg)
            if score > best_score:
                best_score = score
                best_abdc_key = abdc_key
        if best_score >= FUZZY_THRESHOLD and best_abdc_key:
            fuzzy_candidates.append((best_score, best_abdc_key, ajg_key))

    # Sort by score descending and greedily assign (each ABDC/AJG used at most once)
    fuzzy_candidates.sort(key=lambda x: -x[0])
    for score, abdc_key, ajg_key in fuzzy_candidates:
        if abdc_key in matched_abdc_keys or ajg_key in matched_ajg_keys:
            continue
        matched.append((abdc_key, ajg_key))
        matched_abdc_keys.add(abdc_key)
        matched_ajg_keys.add(ajg_key)

    p3_count = len(matched) - p1_count - p2_count

    unmatched_abdc = set(abdc_data.keys()) - matched_abdc_keys
    unmatched_ajg = set(ajg_data.keys()) - matched_ajg_keys

    print(f"  Pass 1 (exact):         {p1_count} matches")
    print(f"  Pass 2 (deep-norm):     {p2_count} matches")
    print(f"  Pass 3 (fuzzy >=0.82):  {p3_count} matches")
    print(f"  Total matched pairs:    {len(matched)}")
    print(f"  ABDC-only journals:     {len(unmatched_abdc)}")
    print(f"  AJG-only journals:      {len(unmatched_ajg)}")

    return matched, unmatched_abdc, unmatched_ajg


# ----------------------------------------------------------------
# Main
# ----------------------------------------------------------------

def main():
    print("Loading ABDC data...")
    abdc_data = load_abdc()
    print(f"  -> {len(abdc_data)} journals")

    print("Loading AJG data...")
    ajg_data = load_ajg()
    print(f"  -> {len(ajg_data)} journals")

    print("Loading FSB points table...")
    points_table = load_fsb_points()
    print(f"  -> {len(points_table)} rating combinations")

    print("Matching journals across sources...")
    matched, unmatched_abdc, unmatched_ajg = match_journals(abdc_data, ajg_data)

    # Build merged journal list
    merged = []
    seen_titles = set()

    # 1. Matched pairs — combine data from both sources
    for abdc_key, ajg_key in matched:
        abdc_entry = abdc_data[abdc_key]
        ajg_entry = ajg_data[ajg_key]

        # Prefer the shorter/cleaner title (usually ABDC), but keep both for reference
        title = abdc_entry["title"]
        publisher = abdc_entry.get("publisher") or ajg_entry.get("publisher", "")
        field = ajg_entry.get("field", "")
        abdc_rating = abdc_entry.get("abdc", "")
        ajg_rating = ajg_entry.get("ajg", "")
        points = calc_points(abdc_rating, ajg_rating, points_table)

        merged.append({
            "title": title,
            "publisher": publisher,
            "field": field,
            "abdc": abdc_rating,
            "ajg": ajg_rating,
            "points": points,
        })
        seen_titles.add(title.lower())

    # 2. ABDC-only journals
    for abdc_key in sorted(unmatched_abdc):
        entry = abdc_data[abdc_key]
        abdc_rating = entry.get("abdc", "")
        points = calc_points(abdc_rating, "", points_table)
        merged.append({
            "title": entry["title"],
            "publisher": entry.get("publisher", ""),
            "field": "",
            "abdc": abdc_rating,
            "ajg": "",
            "points": points,
        })

    # 3. AJG-only journals
    for ajg_key in sorted(unmatched_ajg):
        entry = ajg_data[ajg_key]
        ajg_rating = entry.get("ajg", "")
        points = calc_points("", ajg_rating, points_table)
        merged.append({
            "title": entry["title"],
            "publisher": entry.get("publisher", ""),
            "field": entry.get("field", ""),
            "abdc": "",
            "ajg": ajg_rating,
            "points": points,
        })

    # Sort final list alphabetically
    merged.sort(key=lambda j: j["title"].lower())

    OUTPUT.write_text(json.dumps(merged, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {len(merged)} journals to {OUTPUT}")


if __name__ == "__main__":
    main()
